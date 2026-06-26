from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
import mysql.connector
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import date, datetime
from flask import send_file
from io import BytesIO
import qrcode
from flask import request

app = Flask(__name__)
app.secret_key = "codestock_tcc_secret_key_2024"


def conectar_db():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="root",
            database="codestock"
        )
        return conn
    except Exception as e:
        print("Erro ao conectar ao banco:", e)
        return None


def limpar_cnpj(cnpj):
    return "".join(filter(str.isdigit, cnpj or ""))


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "usuario_id" not in session:
            flash("Faça login para continuar.", "erro")
            return redirect(url_for("login", next=request.url))
        return f(*args, **kwargs)
    return decorated


def perfil_required(*perfis_permitidos):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            perfil_usuario = session.get("usuario_perfil")

            if perfil_usuario not in perfis_permitidos:
                flash("Você não tem permissão para acessar esta página.", "erro")
                return redirect(url_for("home"))

            return f(*args, **kwargs)
        return decorated
    return decorator


def admin_required(f):
    return perfil_required("admin")(f)


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/cadastro", methods=["GET", "POST"])
def cadastro():
    if request.method == "GET":
        return render_template("cadastro.html")

    razao = request.form.get("razao", "").strip()
    fantasia = request.form.get("fantasia", "").strip()
    cnpj = request.form.get("cnpj", "").strip()
    segmento = request.form.get("segmento", "").strip()
    telefone = request.form.get("telefone", "").strip()
    email_empresa = request.form.get("emailEmpresa", "").strip().lower()
    endereco = request.form.get("endereco", "").strip()

    nome_admin = request.form.get("nomeAdmin", "").strip()
    cargo_admin = request.form.get("cargoAdmin", "").strip()
    email_admin = request.form.get("emailAdmin", "").strip().lower()
    matricula = request.form.get("matriculaAdmin", "").strip() or "ADM001"

    senha_empresa = request.form.get("senha", "")
    confirmar_empresa = request.form.get("confirmarSenha", "")
    senha_admin = request.form.get("senhaAdmin", "")
    confirmar_admin = request.form.get("confirmarSenhaAdmin", "")
    termos = request.form.get("termos")

    if not razao or not cnpj or not segmento or not email_empresa or not nome_admin or not email_admin:
        flash("Preencha todos os campos obrigatórios.", "erro")
        return render_template("cadastro.html")

    if not senha_empresa or not confirmar_empresa:
        flash("Informe e confirme a senha empresarial.", "erro")
        return render_template("cadastro.html")

    if not senha_admin or not confirmar_admin:
        flash("Informe e confirme a senha do administrador.", "erro")
        return render_template("cadastro.html")

    if senha_empresa != confirmar_empresa:
        flash("As senhas da empresa não coincidem.", "erro")
        return render_template("cadastro.html")

    if senha_admin != confirmar_admin:
        flash("As senhas do administrador não coincidem.", "erro")
        return render_template("cadastro.html")

    if len(senha_empresa) < 8:
        flash("A senha empresarial deve ter no mínimo 8 caracteres.", "erro")
        return render_template("cadastro.html")

    if len(senha_admin) < 8:
        flash("A senha do administrador deve ter no mínimo 8 caracteres.", "erro")
        return render_template("cadastro.html")

    if not termos:
        flash("Você precisa aceitar os termos para concluir o cadastro.", "erro")
        return render_template("cadastro.html")

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template("cadastro.html")

    cursor = conn.cursor()
    try:
        senha_empresa_hash = generate_password_hash(senha_empresa)
        senha_admin_hash = generate_password_hash(senha_admin)

        cursor.execute("""
            INSERT INTO empresas
                (razao_social, nome_fantasia, cnpj, segmento, telefone, email, endereco, senha_hash)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            razao,
            fantasia,
            limpar_cnpj(cnpj),
            segmento,
            telefone,
            email_empresa,
            endereco,
            senha_empresa_hash
        ))

        empresa_id = cursor.lastrowid

        cursor.execute("""
            INSERT INTO usuarios
                (empresa_id, nome, matricula, cargo, email, perfil, senha_hash)
            VALUES
                (%s, %s, %s, %s, %s, 'admin', %s)
        """, (
            empresa_id,
            nome_admin,
            matricula,
            cargo_admin,
            email_admin,
            senha_admin_hash
        ))

        conn.commit()
        flash(f"Empresa cadastrada com sucesso! Use a matrícula {matricula} e a senha do administrador para acessar.", "sucesso")
        return redirect(url_for("login"))

    except mysql.connector.IntegrityError:
        conn.rollback()
        flash("CNPJ, e-mail empresarial ou matrícula já cadastrado no sistema.", "erro")
        return render_template("cadastro.html")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao cadastrar: {e}", "erro")
        return render_template("cadastro.html")
    finally:
        cursor.close()
        conn.close()


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html", etapa="empresa")

    etapa = request.form.get("etapa", "empresa")

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template("login.html", etapa="empresa")

    cursor = conn.cursor(dictionary=True)

    try:
        if etapa == "empresa":
            cnpj = request.form.get("cnpj", "").strip()
            email_empresa = request.form.get("emailEmpresa", "").strip().lower()
            senha_empresa = request.form.get("senhaEmpresa", "")

            cursor.execute(
                "SELECT * FROM empresas WHERE LOWER(email) = %s",
                (email_empresa,)
            )
            empresa = cursor.fetchone()

            if not empresa:
                flash("E-mail empresarial não encontrado.", "erro")
                return render_template("login.html", etapa="empresa")

            if limpar_cnpj(empresa["cnpj"]) != limpar_cnpj(cnpj):
                flash("CNPJ incorreto.", "erro")
                return render_template("login.html", etapa="empresa")

            if not check_password_hash(empresa["senha_hash"], senha_empresa):
                flash("Senha empresarial incorreta.", "erro")
                return render_template("login.html", etapa="empresa")

            empresa_nome = empresa["nome_fantasia"] or empresa["razao_social"]

            return render_template(
                "login.html",
                etapa="usuario",
                empresa_id=empresa["id"],
                empresa_nome=empresa_nome
            )

        if etapa == "usuario":
            empresa_id = request.form.get("empresa_id", "").strip()
            empresa_nome = request.form.get("empresa_nome", "").strip()
            usuario_digitado = request.form.get("usuario", "").strip()
            senha_usuario = request.form.get("senhaUsuario", "")
            perfil_selecionado = request.form.get("perfil", "").strip()

            if not empresa_id:
                flash("Empresa não identificada. Faça a primeira etapa novamente.", "erro")
                return render_template("login.html", etapa="empresa")

            if not perfil_selecionado:
                flash("Selecione o perfil de acesso antes de continuar.", "erro")
                return render_template(
                    "login.html",
                    etapa="usuario",
                    empresa_id=empresa_id,
                    empresa_nome=empresa_nome
                )

            cursor.execute("""
                SELECT *
                FROM usuarios
                WHERE empresa_id = %s
                  AND (
                    matricula = %s
                    OR email = %s
                    OR nome = %s
                  )
            """, (
                empresa_id,
                usuario_digitado,
                usuario_digitado.lower(),
                usuario_digitado
            ))

            usuario = cursor.fetchone()

            if not usuario:
                flash("Usuário não encontrado. Use a matrícula cadastrada, por exemplo ADM001.", "erro")
                return render_template(
                    "login.html",
                    etapa="usuario",
                    empresa_id=empresa_id,
                    empresa_nome=empresa_nome
                )

            if usuario["perfil"] != perfil_selecionado:
                flash("O perfil selecionado não corresponde ao usuário informado.", "erro")
                return render_template(
                    "login.html",
                    etapa="usuario",
                    empresa_id=empresa_id,
                    empresa_nome=empresa_nome
                )

            if not check_password_hash(usuario["senha_hash"], senha_usuario):
                flash("Senha do usuário incorreta.", "erro")
                return render_template(
                    "login.html",
                    etapa="usuario",
                    empresa_id=empresa_id,
                    empresa_nome=empresa_nome
                )

            session.clear()
            session["usuario_id"] = usuario["id"]
            session["usuario_nome"] = usuario["nome"]
            session["usuario_perfil"] = usuario["perfil"]
            session["empresa_id"] = usuario["empresa_id"]
            session["empresa_nome"] = empresa_nome
            session["login_origem"] = "empresa"

            return redirect(url_for("home"))

        flash("Etapa de login inválida.", "erro")
        return render_template("login.html", etapa="empresa")

    except Exception as e:
        flash(f"Erro no login: {e}", "erro")
        return render_template("login.html", etapa="empresa")
    finally:
        cursor.close()
        conn.close()


@app.route("/login-funcionario", methods=["GET", "POST"])
def login_funcionario():
    if request.method == "GET":
        return render_template("login_funcionario.html")

    cnpj = request.form.get("cnpj", "").strip()
    perfil_selecionado = request.form.get("perfil", "").strip()
    usuario_digitado = request.form.get("usuario", "").strip()
    senha_usuario = request.form.get("senhaUsuario", "")

    if not cnpj or not perfil_selecionado or not usuario_digitado or not senha_usuario:
        flash("Preencha todos os campos para acessar como funcionário.", "erro")
        return render_template("login_funcionario.html")

    perfis_validos = ["admin", "operador", "conferente"]
    if perfil_selecionado not in perfis_validos:
        flash("Perfil de acesso inválido.", "erro")
        return render_template("login_funcionario.html")

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template("login_funcionario.html")

    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute(
            "SELECT id, razao_social, nome_fantasia, cnpj FROM empresas WHERE cnpj = %s",
            (limpar_cnpj(cnpj),)
        )
        empresa = cursor.fetchone()

        if not empresa:
            flash("Empresa não encontrada. Verifique o CNPJ informado.", "erro")
            return render_template("login_funcionario.html")

        cursor.execute("""
            SELECT *
            FROM usuarios
            WHERE empresa_id = %s
              AND (
                matricula = %s
                OR email = %s
                OR nome = %s
              )
        """, (
            empresa["id"],
            usuario_digitado,
            usuario_digitado.lower(),
            usuario_digitado
        ))

        usuario = cursor.fetchone()

        if not usuario:
            flash("Funcionário não encontrado para esta empresa.", "erro")
            return render_template("login_funcionario.html")

        if usuario["perfil"] != perfil_selecionado:
            flash("O perfil selecionado não corresponde ao funcionário informado.", "erro")
            return render_template("login_funcionario.html")

        if not check_password_hash(usuario["senha_hash"], senha_usuario):
            flash("Senha do funcionário incorreta.", "erro")
            return render_template("login_funcionario.html")

        empresa_nome = empresa["nome_fantasia"] or empresa["razao_social"]

        session.clear()
        session["usuario_id"] = usuario["id"]
        session["usuario_nome"] = usuario["nome"]
        session["usuario_perfil"] = usuario["perfil"]
        session["empresa_id"] = usuario["empresa_id"]
        session["empresa_nome"] = empresa_nome
        session["login_origem"] = "funcionario"

        return redirect(url_for("home"))

    except Exception as e:
        flash(f"Erro no login do funcionário: {e}", "erro")
        return render_template("login_funcionario.html")
    finally:
        cursor.close()
        conn.close()

@app.route("/logout")
def logout():
    origem = session.get("login_origem")
    session.clear()

    if origem == "funcionario":
        return redirect(url_for("login_funcionario"))

    return redirect(url_for("index"))


@app.route("/home")
@login_required
def home():
    conn = conectar_db()
    dados = {
        "total_materiais": 0,
        "pendentes": 0,
        "total_etiquetas": 0,
        "total_movimentacoes": 0,
        "materiais_recentes": [],
        "grafico_semana": {
            "Seg": {"entrada": 0, "saida": 0},
            "Ter": {"entrada": 0, "saida": 0},
            "Qua": {"entrada": 0, "saida": 0},
            "Qui": {"entrada": 0, "saida": 0},
            "Sex": {"entrada": 0, "saida": 0}
        }
    }

    if conn:
        cursor = conn.cursor(dictionary=True)
        empresa_id = session["empresa_id"]

        try:
            cursor.execute("SELECT COUNT(*) AS total FROM materiais WHERE empresa_id = %s", (empresa_id,))
            dados["total_materiais"] = cursor.fetchone()["total"]

            cursor.execute(
                "SELECT COUNT(*) AS total FROM materiais WHERE empresa_id = %s AND status = 'Pendente'",
                (empresa_id,)
            )
            dados["pendentes"] = cursor.fetchone()["total"]

            cursor.execute("SELECT COUNT(*) AS total FROM etiquetas WHERE empresa_id = %s", (empresa_id,))
            dados["total_etiquetas"] = cursor.fetchone()["total"]

            cursor.execute("SELECT COUNT(*) AS total FROM movimentacoes WHERE empresa_id = %s", (empresa_id,))
            dados["total_movimentacoes"] = cursor.fetchone()["total"]

            cursor.execute("""
                SELECT codigo, descricao, lote, localizacao, status
                FROM materiais
                WHERE empresa_id = %s
                ORDER BY criado_em DESC
                LIMIT 5
            """, (empresa_id,))
            dados["materiais_recentes"] = cursor.fetchall()

            cursor.execute("""
                SELECT
                    DAYOFWEEK(data) AS dia_semana,
                    tipo,
                    COUNT(*) AS total
                FROM movimentacoes
                WHERE empresa_id = %s
                  AND data >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
                  AND tipo IN ('Entrada', 'Saída')
                GROUP BY dia_semana, tipo
            """, (empresa_id,))

            mapa_dias = {
                2: "Seg",
                3: "Ter",
                4: "Qua",
                5: "Qui",
                6: "Sex"
            }

            for item in cursor.fetchall():
                dia = mapa_dias.get(item["dia_semana"])

                if dia:
                    if item["tipo"] == "Entrada":
                        dados["grafico_semana"][dia]["entrada"] = item["total"]
                    elif item["tipo"] == "Saída":
                        dados["grafico_semana"][dia]["saida"] = item["total"]

        finally:
            cursor.close()
            conn.close()

    altura_grafico = {}

    for dia, valores in dados["grafico_semana"].items():
        altura_grafico[dia] = {
            "entrada": max(valores["entrada"] * 20, 6),
            "saida": max(valores["saida"] * 20, 6)
        }

    dados["altura_grafico"] = altura_grafico

    return render_template("home.html", **dados)


@app.route("/materiais")
@login_required
def materiais():
    conn = conectar_db()
    lista = []

    if conn:
        cursor = conn.cursor(dictionary=True)
        empresa_id = session["empresa_id"]
        try:
            cursor.execute("""
                SELECT id, codigo, descricao, lote, fornecedor, quantidade, unidade,
                       localizacao, data_entrada, status, responsavel, observacao, criado_em
                FROM materiais
                WHERE empresa_id = %s
                ORDER BY criado_em DESC
            """, (empresa_id,))
            lista = cursor.fetchall()
        finally:
            cursor.close()
            conn.close()

    return render_template("materiais.html", materiais=lista)


@app.route("/material/editar/<codigo>", methods=["GET", "POST"])
@login_required
@perfil_required("admin", "operador")
def editar_material(codigo):
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("materiais"))

    empresa_id = session["empresa_id"]
    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT *
            FROM materiais
            WHERE empresa_id = %s AND codigo = %s
        """, (empresa_id, codigo))
        material = cursor.fetchone()

        if not material:
            flash("Material não encontrado.", "erro")
            return redirect(url_for("materiais"))

        if request.method == "GET":
            return render_template("material_editar.html", material=material)

        novo_codigo = request.form.get("codigo", "").strip()
        descricao = request.form.get("descricao", "").strip()
        lote = request.form.get("lote", "").strip()
        fornecedor = request.form.get("fornecedor", "").strip()
        quantidade = request.form.get("quantidade", "0").strip()
        unidade = request.form.get("unidade", "").strip()
        localizacao = request.form.get("localizacao", "").strip()
        status = request.form.get("status", "Pendente").strip()
        data_entrada = request.form.get("dataEntrada", "").strip() or None
        responsavel = request.form.get("responsavel", "").strip()
        observacao = request.form.get("observacao", "").strip()

        status_validos = ["Liberado", "Pendente", "Conferência", "Bloqueado"]

        if not novo_codigo or not descricao or not lote or not fornecedor or not quantidade or not unidade or not localizacao:
            flash("Preencha todos os campos obrigatórios.", "erro")
            return render_template("material_editar.html", material=material)

        if status not in status_validos:
            flash("Status inválido.", "erro")
            return render_template("material_editar.html", material=material)

        try:
            quantidade_float = float(quantidade)
        except ValueError:
            flash("Quantidade inválida.", "erro")
            return render_template("material_editar.html", material=material)

        cursor.execute("""
            UPDATE materiais
            SET codigo = %s,
                descricao = %s,
                lote = %s,
                fornecedor = %s,
                quantidade = %s,
                unidade = %s,
                localizacao = %s,
                status = %s,
                data_entrada = %s,
                responsavel = %s,
                observacao = %s
            WHERE empresa_id = %s AND codigo = %s
        """, (
            novo_codigo,
            descricao,
            lote,
            fornecedor,
            quantidade_float,
            unidade,
            localizacao,
            status,
            data_entrada,
            responsavel,
            observacao,
            empresa_id,
            codigo
        ))

        if novo_codigo != codigo:
            cursor.execute("""
                UPDATE movimentacoes
                SET material_codigo = %s
                WHERE empresa_id = %s AND material_codigo = %s
            """, (novo_codigo, empresa_id, codigo))

            cursor.execute("""
                UPDATE etiquetas
                SET material_codigo = %s,
                    lote = %s,
                    descricao = %s,
                    quantidade = %s,
                    localizacao = %s
                WHERE empresa_id = %s AND material_codigo = %s
            """, (
                novo_codigo,
                lote,
                descricao,
                f"{quantidade_float:g} {unidade}",
                localizacao,
                empresa_id,
                codigo
            ))
        else:
            cursor.execute("""
                UPDATE etiquetas
                SET lote = %s,
                    descricao = %s,
                    quantidade = %s,
                    localizacao = %s
                WHERE empresa_id = %s AND material_codigo = %s
            """, (
                lote,
                descricao,
                f"{quantidade_float:g} {unidade}",
                localizacao,
                empresa_id,
                codigo
            ))

        conn.commit()
        flash("Material atualizado com sucesso!", "sucesso")
        return redirect(url_for("materiais"))

    except mysql.connector.IntegrityError:
        conn.rollback()
        flash("Já existe outro material com esse código nesta empresa.", "erro")
        return redirect(url_for("materiais"))
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao editar material: {e}", "erro")
        return redirect(url_for("materiais"))
    finally:
        cursor.close()
        conn.close()


@app.route("/material/excluir/<codigo>", methods=["POST"])
@login_required
@perfil_required("admin")
def excluir_material(codigo):
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("materiais"))

    empresa_id = session["empresa_id"]
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT id
            FROM materiais
            WHERE empresa_id = %s AND codigo = %s
        """, (empresa_id, codigo))

        if not cursor.fetchone():
            flash("Material não encontrado.", "erro")
            return redirect(url_for("materiais"))

        cursor.execute("""
            DELETE FROM etiquetas
            WHERE empresa_id = %s AND material_codigo = %s
        """, (empresa_id, codigo))

        cursor.execute("""
            DELETE FROM materiais
            WHERE empresa_id = %s AND codigo = %s
        """, (empresa_id, codigo))

        conn.commit()
        flash("Material excluído com sucesso!", "sucesso")

    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir material: {e}", "erro")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("materiais"))


@app.route("/materialcadastro", methods=["GET", "POST"])
@login_required
@perfil_required("admin", "operador")
def materialcadastro():
    if request.method == "GET":
        return render_template("materialcadastro.html")

    codigo = request.form.get("codigo", "").strip()
    descricao = request.form.get("descricao", "").strip()
    lote = request.form.get("lote", "").strip()
    fornecedor = request.form.get("fornecedor", "").strip()
    quantidade = request.form.get("quantidade", "0")
    unidade = request.form.get("unidade", "").strip()
    localizacao = request.form.get("localizacao", "").strip()
    status = request.form.get("status", "Pendente").strip()
    data_entrada = request.form.get("dataEntrada", str(date.today()))
    responsavel = request.form.get("responsavel", "").strip()
    observacao = request.form.get("observacao", "").strip()

    if not codigo or not descricao or not lote or not fornecedor or not quantidade or not unidade or not localizacao:
        flash("Preencha todos os campos obrigatórios.", "erro")
        return render_template("materialcadastro.html")

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template("materialcadastro.html")

    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO materiais
                (empresa_id, codigo, descricao, lote, fornecedor, quantidade, unidade,
                 localizacao, status, data_entrada, responsavel, observacao)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            session["empresa_id"],
            codigo,
            descricao,
            lote,
            fornecedor,
            float(quantidade),
            unidade,
            localizacao,
            status,
            data_entrada,
            responsavel,
            observacao
        ))

        conn.commit()
        flash("Material cadastrado com sucesso!", "sucesso")
        return redirect(url_for("materiais"))

    except mysql.connector.IntegrityError:
        flash(f"O código '{codigo}' já existe para esta empresa.", "erro")
        return render_template("materialcadastro.html")
    except Exception as e:
        flash(f"Erro ao salvar: {e}", "erro")
        return render_template("materialcadastro.html")
    finally:
        cursor.close()
        conn.close()


@app.route("/movimentacoes", methods=["GET", "POST"])
@login_required
@perfil_required("admin", "operador", "conferente")
def movimentacoes():
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template(
            "movimentacoes.html",
            movimentacoes=[],
            materiais=[],
            filtros={},
            resumo={
                "entradas_hoje": 0,
                "saidas_hoje": 0,
                "transferencias_hoje": 0,
                "bloqueios_hoje": 0
            },
            altura_grafico={
                "Seg": {"entrada": 6, "saida": 6},
                "Ter": {"entrada": 6, "saida": 6},
                "Qua": {"entrada": 6, "saida": 6},
                "Qui": {"entrada": 6, "saida": 6},
                "Sex": {"entrada": 6, "saida": 6}
            }
        )

    empresa_id = session["empresa_id"]

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        tipo = request.form.get("tipo", "").strip()
        origem = request.form.get("origem", "").strip()
        destino = request.form.get("destino", "").strip()
        quantidade = request.form.get("quantidade", "0").strip()
        data_mov = request.form.get("data", str(date.today())).strip()
        responsavel = session.get("usuario_nome", "Sistema")
        validacao = request.form.get("validacao", "Validado").strip()
        observacao = request.form.get("observacao", "").strip()

        tipos_validos = [
            "Entrada",
            "Saída",
            "Transferência",
            "Correção de localização",
            "Bloqueio de material"
        ]

        validacoes_validas = ["Validado", "Pendente", "Revisão necessária"]

        if not codigo or not tipo or not quantidade:
            flash("Informe o material, o tipo de movimentação e a quantidade.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        if tipo not in tipos_validos:
            flash("Tipo de movimentação inválido.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        if validacao not in validacoes_validas:
            flash("Status de validação inválido.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        try:
            qtd = float(quantidade.replace(",", "."))
        except ValueError:
            flash("Quantidade inválida.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        if qtd <= 0:
            flash("A quantidade deve ser maior que zero.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        if tipo in ["Entrada", "Transferência", "Correção de localização"] and not destino:
            flash("Informe o local de destino para este tipo de movimentação.", "erro")
            conn.close()
            return redirect(url_for("movimentacoes"))

        cursor = conn.cursor(dictionary=True)

        try:
            cursor.execute("""
                SELECT codigo, descricao, quantidade, localizacao, status
                FROM materiais
                WHERE empresa_id = %s AND codigo = %s
            """, (empresa_id, codigo))
            material = cursor.fetchone()

            if not material:
                flash("Material não encontrado. Verifique o código informado.", "erro")
                return redirect(url_for("movimentacoes"))

            local_atual = (material.get("localizacao") or "").strip()

            if tipo in ["Saída", "Transferência", "Correção de localização"]:
                if not origem:
                    flash(f"Informe a origem. O material está atualmente em: {local_atual or 'sem localização cadastrada'}", "erro")
                    return redirect(url_for("movimentacoes"))

                if origem.lower() != local_atual.lower():
                    flash(f"Local de origem incorreto. O material está atualmente em: {local_atual}", "erro")
                    return redirect(url_for("movimentacoes"))

            if tipo == "Saída" and float(material["quantidade"] or 0) < qtd:
                flash("Estoque insuficiente para realizar a saída.", "erro")
                return redirect(url_for("movimentacoes"))

            if tipo == "Entrada" and not origem:
                origem = "Entrada externa"

            if tipo == "Saída" and not destino:
                destino = "Saída do estoque"

            if tipo == "Bloqueio de material":
                origem = origem or local_atual
                destino = destino or local_atual

            cursor.execute("""
                INSERT INTO movimentacoes
                    (empresa_id, material_codigo, tipo, origem, destino,
                     quantidade, data, responsavel, validacao, observacao)
                VALUES
                    (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                empresa_id,
                codigo,
                tipo,
                origem,
                destino,
                qtd,
                data_mov,
                responsavel,
                validacao,
                observacao
            ))

            if tipo == "Entrada":
                cursor.execute("""
                    UPDATE materiais
                    SET quantidade = quantidade + %s,
                        localizacao = %s
                    WHERE codigo = %s AND empresa_id = %s
                """, (qtd, destino, codigo, empresa_id))

            elif tipo == "Saída":
                cursor.execute("""
                    UPDATE materiais
                    SET quantidade = quantidade - %s
                    WHERE codigo = %s AND empresa_id = %s
                """, (qtd, codigo, empresa_id))

            elif tipo in ["Transferência", "Correção de localização"]:
                cursor.execute("""
                    UPDATE materiais
                    SET localizacao = %s
                    WHERE codigo = %s AND empresa_id = %s
                """, (destino, codigo, empresa_id))

            elif tipo == "Bloqueio de material":
                cursor.execute("""
                    UPDATE materiais
                    SET status = 'Bloqueado'
                    WHERE codigo = %s AND empresa_id = %s
                """, (codigo, empresa_id))

            conn.commit()
            flash("Movimentação registrada com sucesso!", "sucesso")

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao registrar movimentação: {e}", "erro")
        finally:
            cursor.close()

        conn.close()
        return redirect(url_for("movimentacoes"))

    filtros = {
        "busca": request.args.get("busca", "").strip(),
        "tipo": request.args.get("tipo", "").strip(),
        "data_inicio": request.args.get("data_inicio", "").strip(),
        "data_fim": request.args.get("data_fim", "").strip()
    }

    lista = []
    materiais_lista = []
    resumo = {
        "entradas_hoje": 0,
        "saidas_hoje": 0,
        "transferencias_hoje": 0,
        "bloqueios_hoje": 0
    }

    grafico_semana = {
        "Seg": {"entrada": 0, "saida": 0},
        "Ter": {"entrada": 0, "saida": 0},
        "Qua": {"entrada": 0, "saida": 0},
        "Qui": {"entrada": 0, "saida": 0},
        "Sex": {"entrada": 0, "saida": 0}
    }

    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT codigo, descricao, quantidade, unidade, localizacao, status
            FROM materiais
            WHERE empresa_id = %s
            ORDER BY codigo
        """, (empresa_id,))
        materiais_lista = cursor.fetchall()

        cursor.execute("""
            SELECT
                SUM(CASE WHEN tipo = 'Entrada' THEN 1 ELSE 0 END) AS entradas_hoje,
                SUM(CASE WHEN tipo = 'Saída' THEN 1 ELSE 0 END) AS saidas_hoje,
                SUM(CASE WHEN tipo = 'Transferência' THEN 1 ELSE 0 END) AS transferencias_hoje,
                SUM(CASE WHEN tipo = 'Bloqueio de material' THEN 1 ELSE 0 END) AS bloqueios_hoje
            FROM movimentacoes
            WHERE empresa_id = %s
            AND data >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
        """, (empresa_id,))
        resumo_db = cursor.fetchone() or {}
        resumo = {
            "entradas_hoje": resumo_db.get("entradas_hoje") or 0,
            "saidas_hoje": resumo_db.get("saidas_hoje") or 0,
            "transferencias_hoje": resumo_db.get("transferencias_hoje") or 0,
            "bloqueios_hoje": resumo_db.get("bloqueios_hoje") or 0
        }

        cursor.execute("""
            SELECT
                DAYOFWEEK(data) AS dia_semana,
                tipo,
                COUNT(*) AS total
            FROM movimentacoes
            WHERE empresa_id = %s
              AND data >= DATE_SUB(CURDATE(), INTERVAL 7 DAY)
              AND tipo IN ('Entrada', 'Saída')
            GROUP BY dia_semana, tipo
        """, (empresa_id,))

        mapa_dias = {2: "Seg", 3: "Ter", 4: "Qua", 5: "Qui", 6: "Sex"}

        for item in cursor.fetchall():
            dia = mapa_dias.get(item["dia_semana"])
            if dia:
                if item["tipo"] == "Entrada":
                    grafico_semana[dia]["entrada"] = item["total"]
                elif item["tipo"] == "Saída":
                    grafico_semana[dia]["saida"] = item["total"]

        sql = """
            SELECT m.tipo, m.material_codigo, mat.descricao AS material_descricao,
                   m.origem, m.destino, m.quantidade, m.validacao, m.data,
                   m.responsavel, m.observacao, m.criado_em
            FROM movimentacoes m
            LEFT JOIN materiais mat
              ON mat.empresa_id = m.empresa_id
             AND mat.codigo = m.material_codigo
            WHERE m.empresa_id = %s
        """
        params = [empresa_id]

        if filtros["busca"]:
            sql += """
                AND (
                    m.material_codigo LIKE %s
                    OR mat.descricao LIKE %s
                    OR m.responsavel LIKE %s
                    OR m.origem LIKE %s
                    OR m.destino LIKE %s
                )
            """
            termo = f"%{filtros['busca']}%"
            params.extend([termo, termo, termo, termo, termo])

        if filtros["tipo"]:
            sql += " AND m.tipo = %s"
            params.append(filtros["tipo"])

        if filtros["data_inicio"]:
            sql += " AND m.data >= %s"
            params.append(filtros["data_inicio"])

        if filtros["data_fim"]:
            sql += " AND m.data <= %s"
            params.append(filtros["data_fim"])

        sql += " ORDER BY m.criado_em DESC LIMIT 100"

        cursor.execute(sql, params)
        lista = cursor.fetchall()

    finally:
        cursor.close()
        conn.close()

    altura_grafico = {}
    for dia, valores in grafico_semana.items():
        altura_grafico[dia] = {
            "entrada": max(valores["entrada"] * 20, 6),
            "saida": max(valores["saida"] * 20, 6)
        }

    return render_template(
        "movimentacoes.html",
        movimentacoes=lista,
        materiais=materiais_lista,
        filtros=filtros,
        resumo=resumo,
        altura_grafico=altura_grafico
    )


@app.route("/etiquetas", methods=["GET", "POST"])
@login_required
@perfil_required("admin", "operador")
def etiquetas():
    conn = conectar_db()
    etiquetas_lista = []
    materiais_lista = []
    resumo_etiquetas = {
        "total": 0,
        "ativas": 0,
        "pendentes": 0,
        "reimpressoes": 0
    }

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template(
            "etiquetas.html",
            etiquetas=[],
            materiais=[],
            resumo_etiquetas=resumo_etiquetas
        )

    empresa_id = session["empresa_id"]

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip()
        lote = request.form.get("lote", "").strip()
        descricao = request.form.get("descricao", "").strip()
        quantidade = request.form.get("quantidade", "").strip()
        localizacao = request.form.get("localizacao", "").strip()
        status = request.form.get("statusEtiqueta", "Gerada").strip()

        status_validos = ["Ativa", "Gerada", "Pendente", "Reimpressão"]

        if not codigo or not descricao or not quantidade or not localizacao:
            flash("Selecione um material e confirme os dados da etiqueta.", "erro")
            conn.close()
            return redirect(url_for("etiquetas"))

        if status not in status_validos:
            flash("Status da etiqueta inválido.", "erro")
            conn.close()
            return redirect(url_for("etiquetas"))

        cursor = conn.cursor(dictionary=True)
        try:
            cursor.execute("""
                SELECT codigo
                FROM materiais
                WHERE empresa_id = %s AND codigo = %s
            """, (empresa_id, codigo))
            material = cursor.fetchone()

            if not material:
                flash("Material não encontrado para gerar etiqueta.", "erro")
                return redirect(url_for("etiquetas"))

            cursor.execute("""
                INSERT INTO etiquetas
                    (empresa_id, material_codigo, lote, descricao, quantidade, localizacao, status)
                VALUES
                    (%s, %s, %s, %s, %s, %s, %s)
            """, (empresa_id, codigo, lote, descricao, quantidade, localizacao, status))

            conn.commit()
            flash("Etiqueta gerada com sucesso! Ela já está pronta para impressão ou PDF.", "sucesso")
            return redirect(url_for("etiquetas"))

        except Exception as e:
            conn.rollback()
            flash(f"Erro ao gerar etiqueta: {e}", "erro")
        finally:
            cursor.close()

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT codigo, descricao, lote, quantidade, unidade, localizacao, status
            FROM materiais
            WHERE empresa_id = %s
            ORDER BY criado_em DESC
        """, (empresa_id,))
        materiais_lista = cursor.fetchall()

        cursor.execute("""
            SELECT e.material_codigo, e.descricao, e.lote, e.quantidade,
                   e.localizacao, e.status, e.criado_em, m.unidade
            FROM etiquetas e
            LEFT JOIN materiais m
              ON m.empresa_id = e.empresa_id
             AND m.codigo = e.material_codigo
            WHERE e.empresa_id = %s
            ORDER BY e.criado_em DESC
        """, (empresa_id,))
        etiquetas_lista = cursor.fetchall()

        cursor.execute("""
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'Ativa' THEN 1 ELSE 0 END) AS ativas,
                SUM(CASE WHEN status = 'Pendente' THEN 1 ELSE 0 END) AS pendentes,
                SUM(CASE WHEN status = 'Reimpressão' THEN 1 ELSE 0 END) AS reimpressoes
            FROM etiquetas
            WHERE empresa_id = %s
        """, (empresa_id,))
        resumo_db = cursor.fetchone() or {}
        resumo_etiquetas = {
            "total": resumo_db.get("total") or 0,
            "ativas": resumo_db.get("ativas") or 0,
            "pendentes": resumo_db.get("pendentes") or 0,
            "reimpressoes": resumo_db.get("reimpressoes") or 0
        }

    finally:
        cursor.close()
        conn.close()

    return render_template(
        "etiquetas.html",
        etiquetas=etiquetas_lista,
        materiais=materiais_lista,
        resumo_etiquetas=resumo_etiquetas
    )


@app.route("/etiquetas/pdf/<codigo>")
@login_required
@perfil_required("admin", "operador")
def baixar_etiqueta_pdf(codigo):
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("etiquetas"))

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT m.codigo, m.descricao, m.lote, m.quantidade, m.unidade,
                   m.localizacao, m.status, e.criado_em AS etiqueta_criada_em,
                   e.status AS status_etiqueta
            FROM materiais m
            LEFT JOIN etiquetas e
              ON e.empresa_id = m.empresa_id
             AND e.material_codigo = m.codigo
            WHERE m.empresa_id = %s AND m.codigo = %s
            ORDER BY e.criado_em DESC
            LIMIT 1
        """, (session["empresa_id"], codigo))
        material = cursor.fetchone()
    finally:
        cursor.close()
        conn.close()

    if not material:
        flash("Material não encontrado para gerar PDF da etiqueta.", "erro")
        return redirect(url_for("etiquetas"))

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
    except Exception:
        flash("Dependência reportlab não instalada. Rode: pip install reportlab", "erro")
        return redirect(url_for("etiquetas"))

    buffer_pdf = BytesIO()
    pdf = canvas.Canvas(buffer_pdf, pagesize=A4)
    largura, altura = A4

    etiqueta_largura = 100 * mm
    etiqueta_altura = 70 * mm
    x = 18 * mm
    y = altura - etiqueta_altura - 18 * mm

    pdf.setStrokeColor(colors.HexColor("#07182c"))
    pdf.setLineWidth(1.4)
    pdf.roundRect(x, y, etiqueta_largura, etiqueta_altura, 4 * mm, stroke=1, fill=0)

    pdf.setFillColor(colors.HexColor("#07182c"))
    pdf.rect(x, y + etiqueta_altura - 13 * mm, etiqueta_largura, 13 * mm, stroke=0, fill=1)
    pdf.setFillColor(colors.white)
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(x + 6 * mm, y + etiqueta_altura - 8.5 * mm, "CODESTOCK")
    pdf.setFont("Helvetica", 7)
    pdf.drawRightString(x + etiqueta_largura - 6 * mm, y + etiqueta_altura - 8 * mm, "Etiqueta de identificação")

    qr_url = request.host_url + f"material/{material['codigo']}"
    qr_img = qrcode.make(qr_url)
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)
    qr_reader = ImageReader(qr_buffer)

    qr_size = 34 * mm
    qr_x = x + etiqueta_largura - qr_size - 7 * mm
    qr_y = y + 17 * mm
    pdf.drawImage(qr_reader, qr_x, qr_y, qr_size, qr_size, mask="auto")

    info_x = x + 6 * mm
    info_y = y + etiqueta_altura - 22 * mm
    linha = 7 * mm

    quantidade = f"{material.get('quantidade') or 0} {material.get('unidade') or ''}".strip()
    status = material.get("status_etiqueta") or material.get("status") or "Gerada"
    data_geracao = material.get("etiqueta_criada_em") or date.today()

    campos = [
        ("Código", material.get("codigo") or "-"),
        ("Material", material.get("descricao") or "-"),
        ("Lote", material.get("lote") or "-"),
        ("Quantidade", quantidade or "-"),
        ("Local", material.get("localizacao") or "-"),
        ("Status", status),
    ]

    pdf.setFillColor(colors.HexColor("#07182c"))
    for rotulo, valor in campos:
        pdf.setFont("Helvetica-Bold", 7)
        pdf.drawString(info_x, info_y, f"{rotulo}:")
        pdf.setFont("Helvetica", 8)
        texto = str(valor)
        if len(texto) > 34:
            texto = texto[:31] + "..."
        pdf.drawString(info_x + 22 * mm, info_y, texto)
        info_y -= linha

    pdf.setFont("Helvetica", 6.5)
    pdf.setFillColor(colors.HexColor("#5f6b7c"))
    pdf.drawString(x + 6 * mm, y + 7 * mm, f"Gerado em: {data_geracao}")
    pdf.drawRightString(x + etiqueta_largura - 6 * mm, y + 7 * mm, "Escaneie para abrir a ficha do material")

    pdf.showPage()
    pdf.save()
    buffer_pdf.seek(0)

    nome_arquivo = f"etiqueta_{material['codigo']}.pdf"
    return send_file(
        buffer_pdf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=nome_arquivo
    )


@app.route("/usuarios", methods=["GET", "POST"])
@login_required
@perfil_required("admin")
def usuarios():
    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("home"))

    empresa_id = session["empresa_id"]
    perfis_validos = ["admin", "operador", "conferente"]

    if request.method == "POST":
        nome = request.form.get("nome", "").strip()
        matricula = request.form.get("matricula", "").strip()
        cargo = request.form.get("cargo", "").strip()
        email = request.form.get("email", "").strip().lower()
        perfil = request.form.get("perfil", "operador").strip()
        senha = request.form.get("senha", "")
        confirmar_senha = request.form.get("confirmarSenha", "")

        if not nome or not matricula or not perfil or not senha:
            flash("Preencha nome, matrícula, perfil e senha.", "erro")
            conn.close()
            return redirect(url_for("usuarios"))

        if perfil not in perfis_validos:
            flash("Perfil inválido.", "erro")
            conn.close()
            return redirect(url_for("usuarios"))

        if senha != confirmar_senha:
            flash("As senhas do usuário não coincidem.", "erro")
            conn.close()
            return redirect(url_for("usuarios"))

        if len(senha) < 8:
            flash("A senha do usuário deve ter no mínimo 8 caracteres.", "erro")
            conn.close()
            return redirect(url_for("usuarios"))

        cursor = conn.cursor()
        try:
            senha_hash = generate_password_hash(senha)
            cursor.execute("""
                INSERT INTO usuarios
                    (empresa_id, nome, matricula, cargo, email, perfil, senha_hash)
                VALUES
                    (%s, %s, %s, %s, %s, %s, %s)
            """, (
                empresa_id,
                nome,
                matricula,
                cargo,
                email,
                perfil,
                senha_hash
            ))
            conn.commit()
            flash("Usuário cadastrado com sucesso!", "sucesso")
        except mysql.connector.IntegrityError:
            conn.rollback()
            flash("Já existe um usuário com essa matrícula nesta empresa.", "erro")
        except Exception as e:
            conn.rollback()
            flash(f"Erro ao cadastrar usuário: {e}", "erro")
        finally:
            cursor.close()

        conn.close()
        return redirect(url_for("usuarios"))

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT id, nome, matricula, cargo, email, perfil, criado_em
            FROM usuarios
            WHERE empresa_id = %s
            ORDER BY criado_em DESC
        """, (empresa_id,))
        lista_usuarios = cursor.fetchall()

        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE empresa_id = %s", (empresa_id,))
        total_usuarios = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE empresa_id = %s AND perfil = 'admin'", (empresa_id,))
        total_admins = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE empresa_id = %s AND perfil = 'operador'", (empresa_id,))
        total_operadores = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM usuarios WHERE empresa_id = %s AND perfil = 'conferente'", (empresa_id,))
        total_conferentes = cursor.fetchone()["total"]
    finally:
        cursor.close()
        conn.close()

    return render_template(
        "usuarios.html",
        usuarios=lista_usuarios,
        total_usuarios=total_usuarios,
        total_admins=total_admins,
        total_operadores=total_operadores,
        total_conferentes=total_conferentes
    )


@app.route("/usuarios/editar/<int:usuario_id>", methods=["POST"])
@login_required
@perfil_required("admin")
def editar_usuario(usuario_id):
    nome = request.form.get("nome", "").strip()
    matricula = request.form.get("matricula", "").strip()
    cargo = request.form.get("cargo", "").strip()
    email = request.form.get("email", "").strip().lower()
    perfil = request.form.get("perfil", "operador").strip()
    senha = request.form.get("senha", "")
    confirmar_senha = request.form.get("confirmarSenha", "")

    perfis_validos = ["admin", "operador", "conferente"]

    if not nome or not matricula or perfil not in perfis_validos:
        flash("Preencha nome, matrícula e perfil corretamente.", "erro")
        return redirect(url_for("usuarios"))

    if senha and senha != confirmar_senha:
        flash("As novas senhas não coincidem.", "erro")
        return redirect(url_for("usuarios"))

    if senha and len(senha) < 8:
        flash("A nova senha deve ter no mínimo 8 caracteres.", "erro")
        return redirect(url_for("usuarios"))

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("usuarios"))

    cursor = conn.cursor(dictionary=True)
    try:
        cursor.execute(
            "SELECT id, perfil FROM usuarios WHERE id = %s AND empresa_id = %s",
            (usuario_id, session["empresa_id"])
        )
        usuario_atual = cursor.fetchone()

        if not usuario_atual:
            flash("Usuário não encontrado.", "erro")
            return redirect(url_for("usuarios"))

        if usuario_id == session.get("usuario_id") and perfil != "admin":
            flash("Você não pode remover o próprio perfil de administrador.", "erro")
            return redirect(url_for("usuarios"))

        if senha:
            senha_hash = generate_password_hash(senha)
            cursor.execute("""
                UPDATE usuarios
                SET nome = %s, matricula = %s, cargo = %s, email = %s, perfil = %s, senha_hash = %s
                WHERE id = %s AND empresa_id = %s
            """, (
                nome, matricula, cargo, email, perfil, senha_hash, usuario_id, session["empresa_id"]
            ))
        else:
            cursor.execute("""
                UPDATE usuarios
                SET nome = %s, matricula = %s, cargo = %s, email = %s, perfil = %s
                WHERE id = %s AND empresa_id = %s
            """, (
                nome, matricula, cargo, email, perfil, usuario_id, session["empresa_id"]
            ))

        conn.commit()
        flash("Usuário atualizado com sucesso!", "sucesso")
    except mysql.connector.IntegrityError:
        conn.rollback()
        flash("Já existe outro usuário com essa matrícula nesta empresa.", "erro")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao editar usuário: {e}", "erro")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("usuarios"))


@app.route("/usuarios/excluir/<int:usuario_id>", methods=["POST"])
@login_required
@perfil_required("admin")
def excluir_usuario(usuario_id):
    if usuario_id == session.get("usuario_id"):
        flash("Você não pode excluir o próprio usuário logado.", "erro")
        return redirect(url_for("usuarios"))

    conn = conectar_db()
    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("usuarios"))

    cursor = conn.cursor()
    try:
        cursor.execute(
            "DELETE FROM usuarios WHERE id = %s AND empresa_id = %s",
            (usuario_id, session["empresa_id"])
        )
        conn.commit()

        if cursor.rowcount == 0:
            flash("Usuário não encontrado.", "erro")
        else:
            flash("Usuário excluído com sucesso!", "sucesso")
    except Exception as e:
        conn.rollback()
        flash(f"Erro ao excluir usuário: {e}", "erro")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for("usuarios"))




def _normalizar_periodo(periodo, data_inicio, data_fim):
    hoje = date.today()

    if data_inicio:
        inicio = data_inicio
    elif periodo == "7":
        inicio = hoje.replace().isoformat()
        from datetime import timedelta
        inicio = (hoje - timedelta(days=7)).isoformat()
    elif periodo == "30":
        from datetime import timedelta
        inicio = (hoje - timedelta(days=30)).isoformat()
    elif periodo == "90":
        from datetime import timedelta
        inicio = (hoje - timedelta(days=90)).isoformat()
    elif periodo == "ano":
        inicio = date(hoje.year, 1, 1).isoformat()
    else:
        inicio = "1000-01-01"

    fim = data_fim or hoje.isoformat()
    return inicio, fim


def _filtros_relatorios():
    periodo = request.args.get("periodo", "30").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()
    data_inicio, data_fim = _normalizar_periodo(periodo, data_inicio, data_fim)

    return {
        "periodo": periodo,
        "tipo_relatorio": request.args.get("tipo_relatorio", "geral").strip(),
        "data_inicio": data_inicio,
        "data_fim": data_fim,
        "status": request.args.get("status", "").strip(),
        "tipo_movimentacao": request.args.get("tipo_movimentacao", "").strip(),
        "busca": request.args.get("busca", "").strip(),
        "responsavel": request.args.get("responsavel", "").strip()
    }


def _ultimos_6_meses():
    hoje = date.today()
    meses = []
    ano = hoje.year
    mes = hoje.month
    for _ in range(6):
        chave = f"{ano:04d}-{mes:02d}"
        label = f"{mes:02d}/{str(ano)[2:]}"
        meses.append((chave, label))
        mes -= 1
        if mes == 0:
            mes = 12
            ano -= 1
    return list(reversed(meses))


def _buscar_relatorio_dados(conn, empresa_id, filtros):
    cursor = conn.cursor(dictionary=True)
    dados = {
        "resumo": {},
        "grafico_mensal": [],
        "materiais": [],
        "movimentacoes": [],
        "ocorrencias": [],
        "tabela_analitica": [],
        "opcoes": {"status": [], "responsaveis": []},
        "emitido_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "empresa_nome": session.get("empresa_nome", "Empresa"),
        "usuario_nome": session.get("usuario_nome", "Usuário")
    }

    try:
        cursor.execute("SELECT COUNT(*) AS total FROM materiais WHERE empresa_id = %s", (empresa_id,))
        total_materiais = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM etiquetas WHERE empresa_id = %s", (empresa_id,))
        total_etiquetas = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM materiais WHERE empresa_id = %s AND status = 'Pendente'", (empresa_id,))
        total_pendentes = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) AS total FROM materiais WHERE empresa_id = %s AND status = 'Bloqueado'", (empresa_id,))
        total_bloqueados = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT COUNT(*) AS total
            FROM movimentacoes
            WHERE empresa_id = %s AND data BETWEEN %s AND %s
        """, (empresa_id, filtros["data_inicio"], filtros["data_fim"]))
        total_movimentacoes = cursor.fetchone()["total"]

        cursor.execute("""
            SELECT
                SUM(CASE WHEN tipo = 'Entrada' THEN 1 ELSE 0 END) AS entradas,
                SUM(CASE WHEN tipo = 'Saída' THEN 1 ELSE 0 END) AS saidas,
                SUM(CASE WHEN tipo = 'Transferência' THEN 1 ELSE 0 END) AS transferencias,
                SUM(CASE WHEN tipo = 'Bloqueio de material' THEN 1 ELSE 0 END) AS bloqueios
            FROM movimentacoes
            WHERE empresa_id = %s AND data BETWEEN %s AND %s
        """, (empresa_id, filtros["data_inicio"], filtros["data_fim"]))
        mov_resumo = cursor.fetchone() or {}

        dados["resumo"] = {
            "total_materiais": total_materiais,
            "total_etiquetas": total_etiquetas,
            "total_pendentes": total_pendentes,
            "total_bloqueados": total_bloqueados,
            "total_movimentacoes": total_movimentacoes,
            "entradas": mov_resumo.get("entradas") or 0,
            "saidas": mov_resumo.get("saidas") or 0,
            "transferencias": mov_resumo.get("transferencias") or 0,
            "bloqueios": mov_resumo.get("bloqueios") or 0
        }

        meses = _ultimos_6_meses()
        grafico = {chave: {"mes": label, "entrada": 0, "saida": 0, "transferencia": 0} for chave, label in meses}

        cursor.execute("""
            SELECT DATE_FORMAT(data, '%Y-%m') AS mes, tipo, COUNT(*) AS total
            FROM movimentacoes
            WHERE empresa_id = %s
              AND data >= DATE_SUB(CURDATE(), INTERVAL 6 MONTH)
              AND tipo IN ('Entrada', 'Saída', 'Transferência')
            GROUP BY mes, tipo
            ORDER BY mes
        """, (empresa_id,))

        for row in cursor.fetchall():
            chave = row["mes"]
            if chave in grafico:
                if row["tipo"] == "Entrada":
                    grafico[chave]["entrada"] = row["total"]
                elif row["tipo"] == "Saída":
                    grafico[chave]["saida"] = row["total"]
                elif row["tipo"] == "Transferência":
                    grafico[chave]["transferencia"] = row["total"]

        max_valor = max([max(v["entrada"], v["saida"], v["transferencia"]) for v in grafico.values()] + [1])
        dados["grafico_mensal"] = [
            {
                "mes": v["mes"],
                "entrada": max(int((v["entrada"] / max_valor) * 100), 6) if v["entrada"] else 6,
                "saida": max(int((v["saida"] / max_valor) * 100), 6) if v["saida"] else 6,
                "transferencia": max(int((v["transferencia"] / max_valor) * 100), 6) if v["transferencia"] else 6,
                "entrada_total": v["entrada"],
                "saida_total": v["saida"],
                "transferencia_total": v["transferencia"]
            }
            for v in grafico.values()
        ]

        sql_mat = """
            SELECT codigo, descricao, lote, fornecedor, quantidade, unidade, localizacao, status, data_entrada, responsavel
            FROM materiais
            WHERE empresa_id = %s
        """
        params_mat = [empresa_id]

        if filtros["status"]:
            sql_mat += " AND status = %s"
            params_mat.append(filtros["status"])

        if filtros["busca"]:
            sql_mat += """
                AND (
                    codigo LIKE %s OR descricao LIKE %s OR lote LIKE %s OR fornecedor LIKE %s OR localizacao LIKE %s
                )
            """
            termo = f"%{filtros['busca']}%"
            params_mat.extend([termo, termo, termo, termo, termo])

        sql_mat += " ORDER BY criado_em DESC"
        cursor.execute(sql_mat, params_mat)
        dados["materiais"] = cursor.fetchall()

        sql_mov = """
            SELECT m.tipo, m.material_codigo, mat.descricao AS material_descricao, m.origem, m.destino,
                   m.quantidade, m.data, m.responsavel, m.validacao, m.observacao, m.criado_em
            FROM movimentacoes m
            LEFT JOIN materiais mat ON mat.empresa_id = m.empresa_id AND mat.codigo = m.material_codigo
            WHERE m.empresa_id = %s AND m.data BETWEEN %s AND %s
        """
        params_mov = [empresa_id, filtros["data_inicio"], filtros["data_fim"]]

        if filtros["tipo_movimentacao"]:
            sql_mov += " AND m.tipo = %s"
            params_mov.append(filtros["tipo_movimentacao"])

        if filtros["responsavel"]:
            sql_mov += " AND m.responsavel = %s"
            params_mov.append(filtros["responsavel"])

        if filtros["busca"]:
            sql_mov += """
                AND (
                    m.material_codigo LIKE %s OR mat.descricao LIKE %s OR m.origem LIKE %s OR m.destino LIKE %s OR m.responsavel LIKE %s
                )
            """
            termo = f"%{filtros['busca']}%"
            params_mov.extend([termo, termo, termo, termo, termo])

        sql_mov += " ORDER BY m.criado_em DESC"
        cursor.execute(sql_mov, params_mov)
        dados["movimentacoes"] = cursor.fetchall()

        cursor.execute("""
            SELECT DISTINCT responsavel
            FROM movimentacoes
            WHERE empresa_id = %s AND responsavel IS NOT NULL AND responsavel <> ''
            ORDER BY responsavel
        """, (empresa_id,))
        dados["opcoes"]["responsaveis"] = [r["responsavel"] for r in cursor.fetchall()]
        dados["opcoes"]["status"] = ["Liberado", "Pendente", "Conferência", "Bloqueado"]

        dados["tabela_analitica"] = [
            {
                "indicador": "Materiais rastreáveis",
                "setor": "Estoque geral",
                "resultado": f"{total_etiquetas} etiquetas / {total_materiais} materiais",
                "impacto": "Alto",
                "status": "Normal" if total_materiais == 0 or total_etiquetas >= max(1, int(total_materiais * 0.75)) else "Atenção"
            },
            {
                "indicador": "Itens pendentes",
                "setor": "Recebimento",
                "resultado": f"{total_pendentes} registros",
                "impacto": "Médio",
                "status": "Atenção" if total_pendentes else "Normal"
            },
            {
                "indicador": "Materiais bloqueados",
                "setor": "Qualidade",
                "resultado": f"{total_bloqueados} registros",
                "impacto": "Alto" if total_bloqueados else "Baixo",
                "status": "Crítico" if total_bloqueados >= 5 else ("Atenção" if total_bloqueados else "Normal")
            },
            {
                "indicador": "Movimentações no período",
                "setor": "Operação",
                "resultado": f"{total_movimentacoes} registros",
                "impacto": "Médio",
                "status": "Normal"
            }
        ]

        dados["ocorrencias"] = []
        if total_pendentes:
            dados["ocorrencias"].append({"titulo": f"{total_pendentes} material(is) pendente(s)", "descricao": "Itens aguardando liberação, conferência ou regularização."})
        if total_bloqueados:
            dados["ocorrencias"].append({"titulo": f"{total_bloqueados} material(is) bloqueado(s)", "descricao": "Registros com status bloqueado para análise ou controle de qualidade."})
        if not dados["ocorrencias"]:
            dados["ocorrencias"].append({"titulo": "Nenhuma ocorrência crítica", "descricao": "Não há materiais bloqueados ou pendências relevantes no momento."})
        dados["ocorrencias"].append({"titulo": f"{total_movimentacoes} movimentação(ões) no período", "descricao": "Inclui entradas, saídas, transferências e alterações operacionais."})

    finally:
        cursor.close()

    return dados


def _linhas_exportacao(dados, tipo_relatorio):
    if tipo_relatorio == "movimentacoes":
        cabecalho = ["Data", "Tipo", "Código", "Material", "Origem", "Destino", "Quantidade", "Responsável", "Validação"]
        linhas = [
            [
                str(m.get("data") or ""), m.get("tipo") or "", m.get("material_codigo") or "",
                m.get("material_descricao") or "", m.get("origem") or "", m.get("destino") or "",
                float(m.get("quantidade") or 0), m.get("responsavel") or "", m.get("validacao") or ""
            ] for m in dados["movimentacoes"]
        ]
    elif tipo_relatorio == "bloqueados":
        cabecalho = ["Código", "Descrição", "Lote", "Fornecedor", "Quantidade", "Unidade", "Localização", "Status", "Responsável"]
        linhas = [
            [m.get("codigo"), m.get("descricao"), m.get("lote"), m.get("fornecedor"), float(m.get("quantidade") or 0),
             m.get("unidade"), m.get("localizacao"), m.get("status"), m.get("responsavel")]
            for m in dados["materiais"] if m.get("status") == "Bloqueado"
        ]
    else:
        cabecalho = ["Código", "Descrição", "Lote", "Fornecedor", "Quantidade", "Unidade", "Localização", "Status", "Data entrada", "Responsável"]
        linhas = [
            [m.get("codigo"), m.get("descricao"), m.get("lote"), m.get("fornecedor"), float(m.get("quantidade") or 0),
             m.get("unidade"), m.get("localizacao"), m.get("status"), str(m.get("data_entrada") or ""), m.get("responsavel")]
            for m in dados["materiais"]
        ]
    return cabecalho, linhas


@app.route("/relatorios")
@login_required
@perfil_required("admin")
def relatorios():
    filtros = _filtros_relatorios()
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return render_template("relatorios.html", filtros=filtros, dados=None, qs=request.query_string.decode("utf-8"))

    try:
        dados = _buscar_relatorio_dados(conn, session["empresa_id"], filtros)
    finally:
        conn.close()

    return render_template("relatorios.html", filtros=filtros, dados=dados, qs=request.query_string.decode("utf-8"))


@app.route("/relatorios/pdf")
@login_required
@perfil_required("admin")
def relatorios_pdf():
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except Exception:
        flash("Dependência reportlab não instalada. Rode: pip install reportlab", "erro")
        return redirect(url_for("relatorios"))

    filtros = _filtros_relatorios()
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("relatorios"))

    try:
        dados = _buscar_relatorio_dados(conn, session["empresa_id"], filtros)
    finally:
        conn.close()

    cabecalho, linhas = _linhas_exportacao(dados, filtros["tipo_relatorio"])

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm
    )

    styles = getSampleStyleSheet()
    story = []

    tipo_relatorio = filtros.get("tipo_relatorio", "geral")
    if tipo_relatorio == "movimentacoes":
        titulo = "Relatório de Movimentações"
    elif tipo_relatorio == "bloqueados":
        titulo = "Relatório de Materiais Bloqueados"
    else:
        titulo = "Relatório de Estoque"

    story.append(Paragraph("<b>CODESTOCK</b>", styles["Title"]))
    story.append(Paragraph(f"<b>{titulo}</b>", styles["Heading2"]))
    story.append(Paragraph(f"Empresa: <b>{dados['empresa_nome']}</b>", styles["Normal"]))
    story.append(Paragraph(f"Emitido por: <b>{dados['usuario_nome']}</b> em {dados['emitido_em']}", styles["Normal"]))
    story.append(Paragraph(f"Período: {filtros['data_inicio']} até {filtros['data_fim']}", styles["Normal"]))
    story.append(Spacer(1, 6 * mm))

    resumo = dados["resumo"]
    resumo_tabela = [
        ["Materiais", "Movimentações", "Entradas", "Saídas", "Transferências", "Bloqueios"],
        [
            resumo.get("total_materiais", 0),
            resumo.get("total_movimentacoes", 0),
            resumo.get("entradas", 0),
            resumo.get("saidas", 0),
            resumo.get("transferencias", 0),
            resumo.get("bloqueios", 0),
        ]
    ]
    resumo_table = Table(resumo_tabela, repeatRows=1, hAlign="LEFT")
    resumo_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07182C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D8E2F0")),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F7F7F7")),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 8 * mm))

    story.append(Paragraph("<b>Dados analíticos</b>", styles["Heading3"]))

    def celula(valor):
        texto = str(valor if valor is not None else "-")
        return Paragraph(texto, styles["BodyText"])

    dados_tabela = [[celula(col) for col in cabecalho]]
    for linha in linhas:
        dados_tabela.append([celula(valor) for valor in linha])

    if len(dados_tabela) == 1:
        dados_tabela.append([celula("Nenhum registro encontrado")] + [celula("") for _ in cabecalho[1:]])

    largura_util = landscape(A4)[0] - (20 * mm)
    qtd_colunas = max(len(cabecalho), 1)
    col_widths = [largura_util / qtd_colunas for _ in cabecalho]

    tabela = Table(dados_tabela, repeatRows=1, colWidths=col_widths)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F7BE5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8E2F0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tabela)
    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph("Documento gerado automaticamente pelo CodeStock.", styles["Italic"]))

    doc.build(story)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"relatorio_codestock_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    )


@app.route("/relatorios/excel")
@login_required
@perfil_required("admin")
def relatorios_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception:
        flash("Dependência openpyxl não instalada. Rode: pip install openpyxl", "erro")
        return redirect(url_for("relatorios"))

    filtros = _filtros_relatorios()
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("relatorios"))

    try:
        dados = _buscar_relatorio_dados(conn, session["empresa_id"], filtros)
    finally:
        conn.close()

    cabecalho, linhas = _linhas_exportacao(dados, filtros["tipo_relatorio"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório CodeStock"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(6, len(cabecalho)))
    ws.cell(row=1, column=1, value="CODESTOCK - Relatório Operacional")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="07182C")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    ws.append(["Empresa", dados["empresa_nome"]])
    ws.append(["Emitido por", dados["usuario_nome"]])
    ws.append(["Data de emissão", dados["emitido_em"]])
    ws.append(["Período", f"{filtros['data_inicio']} até {filtros['data_fim']}"])
    ws.append([])
    ws.append(cabecalho)

    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F7BE5")
        cell.alignment = Alignment(horizontal="center")

    for linha in linhas:
        ws.append(linha)

    thin = Side(border_style="thin", color="D8E2F0")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max(6, len(cabecalho))):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center")

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 32)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"relatorio_codestock_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )


@app.route("/material/<codigo>")
@login_required
def visualizar_material(codigo):
    conn = conectar_db()

    if not conn:
        flash("Erro ao conectar ao banco de dados.", "erro")
        return redirect(url_for("materiais"))

    cursor = conn.cursor(dictionary=True)

    try:
        cursor.execute("""
            SELECT *
            FROM materiais
            WHERE empresa_id = %s AND codigo = %s
        """, (session["empresa_id"], codigo))

        material = cursor.fetchone()

        if not material:
            flash("Material não encontrado.", "erro")
            return redirect(url_for("materiais"))

        cursor.execute("""
            SELECT tipo, origem, destino, quantidade, data, responsavel, validacao, observacao, criado_em
            FROM movimentacoes
            WHERE empresa_id = %s AND material_codigo = %s
            ORDER BY criado_em DESC
            LIMIT 20
        """, (session["empresa_id"], codigo))

        historico = cursor.fetchall()

        return render_template(
            "material_detalhes.html",
            material=material,
            historico=historico
        )

    finally:
        cursor.close()
        conn.close()


@app.route("/qrcode/material/<codigo>")
@login_required
def gerar_qrcode_material(codigo):
    url_material = request.host_url + f"material/{codigo}"

    img = qrcode.make(url_material)

    buffer = BytesIO()
    img.save(buffer, format="PNG")
    buffer.seek(0)

    return send_file(buffer, mimetype="image/png")

@app.route("/scan")
@login_required
def scan():
    return render_template("scan.html")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)